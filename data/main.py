import json
from openpyxl import load_workbook

def excel_to_json():
    """
    读取1.xlsx文件并将数据转换为JSON字符串
    """
    try:
        # 加载Excel工作簿
        workbook = load_workbook('1.xlsx')
        sheet = workbook.active
        
        # 获取所有数据
        data = []
        
        # 获取表头（第一行）
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # 获取数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
        
        # 转换为JSON字符串，设置缩进为2使输出更美观
        json_string = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        
        return json_string
    
    except FileNotFoundError:
        return "错误：找不到文件 1.xlsx"
    except Exception as e:
        return f"错误：{str(e)}"

if __name__ == "__main__":
    # 转换Excel数据为JSON
    json_result = excel_to_json()
    
    # 打印JSON结果
    print(json_result)
    
    # 将JSON结果保存到txt文件
    try:
        with open('data.txt', 'w', encoding='utf-8') as f:
            f.write(json_result)
        print("\n数据已成功保存到 data.txt 文件中！")
    except Exception as e:
        print(f"\n保存文件时出错：{str(e)}")